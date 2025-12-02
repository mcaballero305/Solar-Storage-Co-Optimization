import os
import pandas as pd
import numpy as np
import argparse
import logging
from amplpy import AMPL, add_to_path
from datetime import timedelta
from openpyxl import load_workbook
#from emissions_credits import configure_emissions
from microgrid.core.model_helpers import (
    safe_int,
    is_on_peak,
    merge_sweep_summaries,
    reset_ampl_model,
    annualized_battery_cost,
    compute_monthly_summary,
    annualized_pv_cost,
    effective_pv_inputs_by_origin,
    effective_battery_inputs_by_origin,
    make_hourly_csv_name,
    append_scalar_to_sweep_summary,
)


from microgrid.core.pv_sizing_limits import compute_pv_sizing_limits

from microgrid.core.logging_utils import (
    log_simulation_start,
    log_simulation_params,
    log_sheet_name,
    log_simulation_complete,
)

from microgrid.core.excel_utils import write_results_to_excel


#from emissions_credits import configure_emissions





# ---------------------------------------------
# Argument Parsing
# ---------------------------------------------


def extract_value(param):
    return param.value() if hasattr(param, "value") else param



# ---------------------------------------------
# Single-run Logic
# ---------------------------------------------
def run_single_pv(
    run_id: str,
    ampl: AMPL,
    data: pd.DataFrame,
    pv_data: pd.DataFrame,
    pv_row: pd.Series,
    bat_data: pd.DataFrame,
    bat_row: pd.Series,
    excel_filename: str,
    override_params: dict = {},
    sizing_limits: dict = None):


    def _insert_small_scalar_top(excel_path: str, sheet_name: str, mini_df: pd.DataFrame):
  
        wb = _safe_load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return
        ws = wb[sheet_name]
        rows_to_insert = len(mini_df) + 2  # +1 header row +1 blank spacer
        ws.insert_rows(1, amount=rows_to_insert)

    # header
        for j, col in enumerate(mini_df.columns, start=1):
            ws.cell(row=1, column=j, value=str(col))
    # values
        for i, (_, row) in enumerate(mini_df.iterrows(), start=2):
            for j, col in enumerate(mini_df.columns, start=1):
                ws.cell(row=i, column=j, value=row[col])

        wb.save(excel_path)
        wb.close()


    
    enforce_limit_flag = bool(safe_int(override_params.get("enforce_limit_flag", 0)))
    enable_sell        = bool(safe_int(override_params.get("enable_sell", 1)))


    try:
        log_simulation_start(run_id, pv_row["PV_types"], bat_row["BATTERY_TYPES"])

        
        # Clean casting of scalar sweep parameters
        itc_pv = float(override_params.get("itc_pv", 0.0))
        itc_bat = float(override_params.get("itc_bat", 0.0))
        offset_price = float(override_params.get("offset_price", 0.0))

        #print(f"🛠 sizing_limits inside run_single_pv: {sizing_limits}")

        max_kw = float(sizing_limits.get("Final Allowed System Size (kW)", 0.0))
        #print(f"✅ Assigned sizing_limit = {max_kw}")

        A_tot = override_params.get("A_tot", 1e6)
        if pd.isna(A_tot):
            A_tot = 1e6
        A_tot = float(A_tot)
        ampl.param["A_tot"] = A_tot

        #print(f"✅ Assigned A_tot = {A_tot}")
        
        
        ampl.param["offset_price"] = offset_price
        #print(f"✅ Assigned offset_price = {offset_price}")

        
        
        # === Enforce E/P ratio range ===
        ampl.param["E_P_min"] = 2
        ampl.param["E_P_max"] = 6

        ampl.param["sizing_limit"] = max_kw
        
        print(f"🔎 About to assign itc_pv in AMPL")
        ampl.param["itc_pv"] = itc_pv
       # print(f"✅ itc_pv assigned")
        
        ampl.param["itc_bat"]      = itc_bat

       # print(f"✅ Assigned itc_pv = {itc_pv}")
       # print(f"✅ Assigned itc_bat  = {itc_bat}")
       # print(f"✅ Assigned sizing_limit = {max_kw}")


        ampl.set["HOURS"] = data["hour_id"].tolist()
        ampl.set["MONTHS"] = list(range(1, 13))
        ampl.set["onpeak"] = data.loc[data["onpeak"], "hour_id"].tolist()
        ampl.set["PV_types"] = [pv_row["PV_types"]]
        ampl.set["BATTERY_TYPES"] = [bat_row["BATTERY_TYPES"]]

        ampl.param["selected_PV"] = pv_row["PV_types"]
        ampl.param["selected_Battery"] = bat_row["BATTERY_TYPES"]

        # Corrected param assignment (removed invalid set())
        ampl.param["A"] = 1.6
        ampl.param["project_life"] = 25
        ampl.param["discount_rate"] = 0.06
        ampl.param["degradation"] = 0.005
        ampl.param["CRF_pv"] = 0.078227
        ampl.param["cost_sell"] = 0.039
        ampl.param["demand_rate"] = 15.2 # new demand w/updated costs
        ampl.param["max_demand_rate"] = 3.1 # new max demand w/updated costs
        ampl.param["monthly_charge"] = 89.26

        ampl.param["pv_cost"].set_values({(pv_row["PV_types"],): pv_row["pv_cost"]})
        ampl.param["STC_eff"].set_values({(pv_row["PV_types"],): pv_row["STC_eff"]})
        ampl.param["temp_coeff"].set_values({(pv_row["PV_types"],): pv_row["temp_coeff"]})
        ampl.param["O_M"].set_values({(pv_row["PV_types"],): pv_row["O_M"]})
        ampl.param["kW_stc"].set_values({(pv_row["PV_types"],): pv_row["kW_stc"]})

        ampl.param["B_rt_eff"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["B_rt_eff"]})
        ampl.param["B_dod"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["B_dod"]})
        ampl.param["B_cost_e"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["B_cost_e"]})
        ampl.param["B_cost_p"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["B_cost_p"]})
        ampl.param["B_power_rating"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["B_power_rating"]})
        ampl.param["B_O_M"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["B_O_M"]})
        ampl.param["E_Life"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["Energy_Life"]})
        ampl.param["P_Life"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["Power_Life"]})
        ampl.param["E_Repl_Frac"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["Replacement_Fraction_E"]})
        ampl.param["P_Repl_Frac"].set_values({(bat_row["BATTERY_TYPES"],): bat_row["Replacement_Fraction_P"]})

        ampl.param["ch_eff"].set_values({(bat_row["BATTERY_TYPES"],): np.sqrt(bat_row["B_rt_eff"])})
        ampl.param["dis_eff"].set_values({(bat_row["BATTERY_TYPES"],): np.sqrt(bat_row["B_rt_eff"])})

             
        
        
        eff_pv_base, eff_itc_pv = effective_pv_inputs_by_origin(
            pv_row=pv_row,
            itc_pv=itc_pv,            # .30/.40/.50 from CSV
            overrides=override_params # contains use_origin_policy, pv_us_list, pv_cn_list, cn_pv_discount_mult
        )

        eff_b_e, eff_b_p, eff_itc_bat = effective_battery_inputs_by_origin(
            bat_row=bat_row,
            itc_bat=itc_bat,          # .30/.40/.50 from CSV
            overrides=override_params # contains bat_us_list, bat_cn_list, cn_bat_discount_mult_e/_p
        )

        ac_e, ac_p = annualized_battery_cost(
            base_cost_e=eff_b_e,
            base_cost_p=eff_b_p,
            B_O_M=float(bat_row["B_O_M"]),
            energy_life=int(bat_row["Energy_Life"]),
            power_life=int(bat_row["Power_Life"]),
            project_life=25,
            discount_rate=0.06,
            replacement_fraction_e=float(bat_row["Replacement_Fraction_E"]),
            replacement_fraction_p=float(bat_row["Replacement_Fraction_P"]),
            itc_bat=eff_itc_bat
        )
        
        # --- PV Annualized Cost (same structure as battery) ---

        ac_pv = annualized_pv_cost(
            base_cost=eff_pv_base,
            o_m=pv_row["O_M"],
            project_life=25,
            discount_rate=0.06,
            itc_pv=eff_itc_pv
        )
               
        ampl.param["ac_e"].set_values({(bat_row["BATTERY_TYPES"],): ac_e})
        ampl.param["ac_p"].set_values({(bat_row["BATTERY_TYPES"],): ac_p})
        ampl.param["ac_pv"].set_values({(pv_row["PV_types"],): ac_pv})


        ampl.param["L"] = dict(zip(data["hour_id"], data["consumption"]))
        ampl.param["I"] = dict(zip(data["hour_id"], data["insolation"]))        
        ampl.param["Temp"] = dict(zip(data["hour_id"], data["temperature"]))
        ampl.param["cost_grid"] = dict(zip(data["hour_id"], data["C_buy"]))
        ampl.param["month_of"] = dict(zip(data["hour_id"], data["month"]))
        ampl.param["EF_grid"] = dict(zip(data["hour_id"], data["Emissions"]))


        hour_ids = data["hour_id"].tolist()

        if 0 not in hour_ids:
            raise ValueError("Hour 0 is required in the dataset for SOC initialization.")

        prev_hours = {h: h - 1 for h in hour_ids if h > 0}
        #prev_hours = {h: hour_ids[i - 1] for i, h in enumerate(hour_ids) if i > 0}
        ampl.param["prev_h"].set_values(prev_hours)

        log_simulation_params(ampl, pv_row["PV_types"], bat_row["BATTERY_TYPES"])

        if not enable_sell:
            gs = ampl.getVariable("Grid_sell")
            try:
                gs.unfix()
            except Exception:
                pass
            if not enable_sell:
                gs.setValues({(h,): 0.0 for h in data["hour_id"]})
                gs.fix()

        ampl.option['solver'] = 'gurobi'
        ampl.setOption('gurobi_options', 'dualreductions=0')
        ampl.solve()
        
        pv_key   = pv_row["PV_types"]
        P_pv_val = ampl.getVariable("P_pv").get(pv_key).value()
        Total_Energy_val = ampl.getVariable("Total_Energy").value()
        Total_Power_val  = ampl.getVariable("Total_Power").value()
        #ep_ratio = Total_Energy_val / (Total_Power_val + 1e-6)
        #print(f"🔎 E/P Ratio for {run_id}: {ep_ratio:.2f} hrs (Energy: {Total_Energy_val:.2f} kWh, Power: {Total_Power_val:.2f} kW)")



        discount_rate = float(extract_value(ampl.param["discount_rate"]))
        project_life  = int(extract_value(ampl.param["project_life"]))
        crf_pv        = float(extract_value(ampl.param["CRF_pv"]))


        # get O&M $/kW‐yr for this module:
        O_M_val = float(ampl.param["O_M"].get((pv_key,)))

        # installed‐cost breakdown
        
        #print(f"🔎 About to use itc_pv in cost calculation")
        #installed_pv_cost = (pv_row["pv_cost"] * (1 - itc_pv) * crf_pv + O_M_val) * P_pv_val
        installed_pv_cost = ac_pv * P_pv_val   
        print(f"✅ installed_pv_cost calculated")
        
        installed_battery_energy_cost = ac_e * Total_Energy_val
        installed_battery_power_cost  = ac_p * Total_Power_val
        installed_battery_cost        = installed_battery_energy_cost + installed_battery_power_cost

        # grid import cost
        hours = data["hour_id"].tolist()
        grid_import_cost = sum(
            ampl.getVariable("Grid_buy").get(h).value()
            * ampl.getParameter("cost_grid").get(h)
            for h in hours
        )
    
        # sale revenue
        sell_price = float(extract_value(ampl.param["cost_sell"]))
        total_sale_revenue = sum(
            ampl.getVariable("Grid_sell").get(h).value() * sell_price
            for h in hours
        )

        #print(f"[{run_id}] Building config string with enforce_limit={enforce_limit_flag}, enable_sell={enable_sell}")

        config_id_full = (
            f"{pv_row['PV_types']}_{bat_row['BATTERY_TYPES']}"
            f"_ITC{itc_pv:.2f}"
            f"_ITC{itc_bat:.2f}"
            f"_Lim{float(enforce_limit_flag):.2f}"
            f"_Sell{float(enable_sell):.2f}"
)       

        pv_type = pv_row["PV_types"].strip()
        # Use first word + first letter of second word if it exists
        if "Bifacial" in pv_type:
            pv_abbrev = "Mono_BF"
        elif "PERC" in pv_type:
            pv_abbrev = "Mono_PERC"
        elif "CdTe" in pv_type:
            pv_abbrev = "CdTe"
        else:
    # fallback if not matching anything known
            pv_abbrev = pv_type.replace(" ", "_")[:15]
        
        bat_type = bat_row["BATTERY_TYPES"].strip()
        bat_abbrev = bat_type[:5]  # first 5 characters

        # Construct the sheet name as: first5PV_first5Battery_parameter_sweepValue
        itc_token = f"pv{int(round(itc_pv*100))}_bat{int(round(itc_bat*100))}"
        area_token = f"A{int(round(A_tot))}"
        generated_sheet = f"{area_token}_{pv_abbrev}_{bat_abbrev}_{run_id}"
        
        hourly_sheet_name = f"H_{generated_sheet}"[:31]
        monthly_sheet_name = f"M_{generated_sheet}"[:31]
        log_sheet_name(run_id, hourly_sheet_name, generated_sheet)


        hourly_results = []
        P_ch_vals = []
        P_dis_vals = []
        for h in data["hour_id"]:
            try:
                pvL   = ampl.getVariable("PV_gen_L").get((h, pv_row['PV_types'])).value()
                pvG   = ampl.getVariable("PV_gen_G").get((h, pv_row['PV_types'])).value()
                bdisL = ampl.getVariable("B_dis_offset").get((h, bat_row['BATTERY_TYPES'])).value()
                bdisG = ampl.getVariable("B_dis_offset_grid").get((h, bat_row['BATTERY_TYPES'])).value()
                ef    = ampl.getParameter("EF_grid").get(h)
                
                P_ch_val = ampl.getVariable("P_ch").get((h, bat_row['BATTERY_TYPES'])).value()
                P_dis_val = ampl.getVariable("P_dis").get((h, bat_row['BATTERY_TYPES'])).value()
                P_ch_vals.append(P_ch_val)
                P_dis_vals.append(P_dis_val)
                
                hourly_results.append({
                    "Hour": h,
                    "Consumption": ampl.getParameter("L").get(h),
                    "Insolation": ampl.getParameter("I").get(h),
                    "Temperature": ampl.getParameter("Temp").get(h),
                    "Total PV Generation": ampl.getVariable("PV_gen").get((h, pv_row['PV_types'])).value(),
                    "PV to Load":  pvL,
                    "PV to Grid":  pvG,
                    "PV to Battery": ampl.getVariable("P_ch_pv").get((h, bat_row['BATTERY_TYPES'])).value(),
                    "Total Battery Charge": P_ch_val,
                    "Charge from Grid": ampl.getVariable("P_ch_grid").get((h, bat_row['BATTERY_TYPES'])).value(),
                    "Total Battery Discharge": P_dis_val,
                    "Battery to Load": ampl.getVariable("P_dis_l").get((h, bat_row['BATTERY_TYPES'])).value(),
                    "Battery to Grid": ampl.getVariable("P_dis_grid").get((h, bat_row['BATTERY_TYPES'])).value(),
                    "SOC": ampl.getVariable("SOC").get((h, bat_row['BATTERY_TYPES'])).value(),
                    "Grid Import": ampl.getVariable("Grid_buy").get(h).value(),
                    "Grid Export": ampl.getVariable("Grid_sell").get(h).value(),
                    "Grid Price": ampl.getParameter("cost_grid").get(h),
                    "EF_grid": ef,
                    "PV_Charged_Discharge_to_Load": ampl.getVariable("B_dis_offset").get((h, bat_row['BATTERY_TYPES'])).value(),
                    "PV_Charged_Discharge_to_Grid": ampl.getVariable("B_dis_offset_grid").get((h, bat_row['BATTERY_TYPES'])).value(),
                    "SOC_renew": ampl.getVariable("SOC_renew").get((h, bat_row['BATTERY_TYPES'])).value(),
                    "Abatement_Energy_kWh": (pvL + pvG + bdisL + bdisG)
                })



            except Exception as e:
                print(f"Error extracting hourly data for hour {h}: {str(e)}")
                continue   

        
        hourly_df = pd.DataFrame(hourly_results)

            # --- On-peak vs Off-peak label (from driver table 'data') ---
        on_map = dict(zip(data["hour_id"], data["onpeak"]))
        hourly_df["On_peak/Off_peak"] = hourly_df["Hour"].map(on_map).map(lambda x: "on_peak" if bool(x) else "off_peak")

# --- Derive Grid to Load (keep your naming) ---
# Grid to Load = Grid Import - Charge from Grid (any tiny negatives clipped to 0)
        hourly_df["Grid to Load"] = (hourly_df["Grid Import"] - hourly_df["Charge from Grid"]).clip(lower=0)

# --- Battery charging shares (per hour) ---
        den_ch = (hourly_df["PV to Battery"] + hourly_df["Charge from Grid"]).replace(0, np.nan)
        hourly_df["%Batt_Charge_from_PV"]   = (100 * hourly_df["PV to Battery"]     / den_ch).round(2)
        hourly_df["%Batt_Charge_from_Grid"] = (100 * hourly_df["Charge from Grid"]   / den_ch).round(2)

# --- Load coverage shares (per hour) ---
        den_ld = hourly_df["Consumption"].replace(0, np.nan)  # you already add "Consumption"
        hourly_df["%Load_from_PV"]      = (100 * hourly_df["PV to Load"]    / den_ld).round(2)
        hourly_df["%Load_from_Battery"] = (100 * hourly_df["Battery to Load"]/ den_ld).round(2)
        hourly_df["%Load_from_Grid"]    = (100 * hourly_df["Grid to Load"]  / den_ld).round(2)




        # Normalize the key once
        if "Abatement_Energy_KWh" in hourly_df.columns and "Abatement_Energy_kWh" not in hourly_df.columns:
            hourly_df.rename(columns={"Abatement_Energy_KWh": "Abatement_Energy_kWh"}, inplace=True)

        # Compute abated emissions (post-append)
        hourly_df["Abated_Emissions"] = hourly_df["Abatement_Energy_kWh"] * hourly_df["EF_grid"]
        
        # price is $/tCO2, emissions are in tCO2 → $/hr
        offset_price_val = float(ampl.param["offset_price"].value())
        hourly_df["Abatement_Value_$"] = hourly_df["Abated_Emissions"] * offset_price_val

        
        # Monthly summary for costs/charges and emissions rollups
        monthly_df = compute_monthly_summary(hourly_df, data, ampl)


        if "Demand_Charge" in monthly_df.columns and "Max_Demand_Charge" in monthly_df.columns:
            demand_charge_total = monthly_df["Demand_Charge"].sum()
            max_demand_charge_total = monthly_df["Max_Demand_Charge"].sum()
        else:
            demand_charge_total = 0
            max_demand_charge_total = 0
            print(f"⚠️ Warning: Demand_Charge or Max_Demand_Charge not found in {hourly_sheet_name}")


        yearly_import_cost = (
            monthly_df["OffPeak_Cost"].sum()
            + monthly_df["OnPeak_Cost"].sum()
        )
        
        switch_count = 0
        threshold = 1e-4
        for i in range(1, len(P_ch_vals)):
            if P_ch_vals[i - 1] > threshold and P_dis_vals[i] > threshold:
                switch_count += 1

# Assign scalar cost values
        pv_nominal_cost = pv_row['pv_cost']
        bat_nominal_cost_e = bat_row['B_cost_e']
        bat_nominal_cost_p = bat_row['B_cost_p']
        pv_annualized_cost = ac_pv
        bat_annualized_cost_e = ac_e
        bat_annualized_cost_p = ac_p
        Total_Emissions_Offset = ampl.getVariable("Total_Emissions_Offset").value()
        gross_emissions = sum(ampl.getVariable("Grid_buy").get(h).value() * ampl.getParameter("EF_grid").get(h)
                            for h in data["hour_id"])
        abatement_credit = offset_price * Total_Emissions_Offset

        # Annual totals from hourly_df
        annual_pv_to_load   = float(hourly_df["PV to Load"].sum())
        annual_batt_to_load = float(hourly_df["Battery to Load"].sum())
        annual_grid_to_load = float(hourly_df["Grid to Load"].sum())
        annual_load_total   = float(hourly_df["Consumption"].sum())

        annual_pv_to_batt   = float(hourly_df["PV to Battery"].sum())
        annual_grid_to_batt = float(hourly_df["Charge from Grid"].sum())
        annual_batt_charge  = annual_pv_to_batt + annual_grid_to_batt

# Shares (guard for divide-by-zero)
        Annual_pct_Load_from_PV      = 100 * annual_pv_to_load   / annual_load_total if annual_load_total > 0 else np.nan
        Annual_pct_Load_from_Battery = 100 * annual_batt_to_load / annual_load_total if annual_load_total > 0 else np.nan
        Annual_pct_Load_from_Grid    = 100 * annual_grid_to_load / annual_load_total if annual_load_total > 0 else np.nan

        Annual_pct_Batt_from_PV      = 100 * annual_pv_to_batt   / annual_batt_charge if annual_batt_charge > 0 else np.nan
        Annual_pct_Batt_from_Grid    = 100 * annual_grid_to_batt / annual_batt_charge if annual_batt_charge > 0 else np.nan



# Print nicely
        print("\n🔎 === Cost Summary for Current Configuration ===")
        print(f"PV Nominal Cost ($/kW):              {pv_nominal_cost:.2f}")
        print(f"Battery Nominal Cost E ($/kWh):       {bat_nominal_cost_e:.2f}")
        print(f"Battery Nominal Cost P ($/kW):        {bat_nominal_cost_p:.2f}")
        print(f"PV Annualized Cost ($/kW-yr):         {pv_annualized_cost:.2f}")
        print(f"Battery Annualized Cost E ($/kWh-yr): {bat_annualized_cost_e:.2f}")
        print(f"Battery Annualized Cost P ($/kW-yr):  {bat_annualized_cost_p:.2f}")
        print("===============================================\n")


# Debug print for verifying cost breakdown against AMPL objective
        #print("\n🔎 Cost Breakdown Debug:")
        #print(f"  Installed PV Cost:            ${installed_pv_cost:,.2f}")
        #print(f"  Installed Battery Cost:       ${installed_battery_cost:,.2f}")
        #print(f"  Grid Import Cost:             ${yearly_import_cost:,.2f}")
        #print(f"  Demand Charges:               ${demand_charge_total + max_demand_charge_total:,.2f}")
        #print(f"  Monthly Fixed Charge (12x):   ${ampl.param['monthly_charge'].value() * 12:,.2f}")
        #print(f"  Sale Revenue (subtract):      ${total_sale_revenue:,.2f}")
        #print(f"  Computed Total (no obj):      ${installed_pv_cost + installed_battery_cost + yearly_import_cost + demand_charge_total + max_demand_charge_total - total_sale_revenue:,.2f}")
        #print(f"  AMPL Objective (Total Cost):  ${ampl.getObjective('Total_Cost').value():,.2f}")
           
        small_scalar_data = {
            "Installed PV Size Limit (kW)": max_kw,
            "Max PV Array Area": int(A_tot),
            "Total Cost ($)": ampl.getObjective("Total_Cost").value(),
            "Switch Count": switch_count,
            "Number of Panels": ampl.getVariable("N").get(pv_row['PV_types']).value(),
            "Installed Area (m^2)": 1.6 * ampl.getVariable("N").get(pv_row['PV_types']).value(),
            "PV Capacity (kW)": ampl.getVariable("P_pv").get(pv_row['PV_types']).value(),
            "Battery Energy Capacity (kWh)": ampl.getVariable("Total_Energy").value(),
            "Battery Power (kW)": ampl.getVariable("Total_Power").value(),
            "Annual PV Generation (kWh)": sum(ampl.getVariable("PV_gen").get((h, pv_row['PV_types'])).value() for h in data["hour_id"]),
            "Annual Grid Import (kWh)": sum(ampl.getVariable("Grid_buy").get(h).value() for h in data["hour_id"]),
            "Annual Grid Export (kWh)": sum(ampl.getVariable("Grid_sell").get(h).value() for h in data["hour_id"]),
            "Annual % Load from PV": round(Annual_pct_Load_from_PV, 2),
            "Annual % Load from Battery": round(Annual_pct_Load_from_Battery, 2),
            "Annual % Load from Grid": round(Annual_pct_Load_from_Grid, 2),
            "Annual % Batt Charge from PV": round(Annual_pct_Batt_from_PV, 2),
            "Annual % Batt Charge from Grid": round(Annual_pct_Batt_from_Grid, 2),
        }
        # build the mini summary row (the thing you want above hourly)
        small_scalar_df = pd.DataFrame([small_scalar_data])


        scalar_data = {
            #0 Flags
            "Run_ID": run_id,
            "Configuration": config_id_full,
            "itc_PV": itc_pv,
            "itc_Bat": itc_bat,
            "Utility Limit Applied?": int(enforce_limit_flag),
            "Enable Sell Flag":  int(enable_sell),
            "Installed PV Size Limit (kW)": max_kw,
            "Max PV Array Area": int(A_tot),
            "Total Cost ($)": ampl.getObjective("Total_Cost").value(),

            # 3. Component Size (amount)
            "Number of Panels": ampl.getVariable("N").get(pv_row['PV_types']).value(),
            "Installed Area (m^2)": 1.6 * ampl.getVariable("N").get(pv_row['PV_types']).value(),
            
            # 4. Component Size (power/energy)
            "PV Capacity (kW)": ampl.getVariable("P_pv").get(pv_row['PV_types']).value(),
            "Battery Energy Capacity (kWh)": ampl.getVariable("Total_Energy").value(),
            "Battery Power (kW)": ampl.getVariable("Total_Power").value(),
            "Battery E/P Ratio (hrs)": Total_Energy_val / (Total_Power_val + 1e-6),


            # 1. Configuration

            "PV Type": pv_row['PV_types'],
            "Battery Type": bat_row['BATTERY_TYPES'],
            "Switch Count": switch_count,
    


            # 2. Nominal Component Costs
            "PV Cost ($/kW)": pv_row['pv_cost'], 
            "Battery Cost E ($/kWh)": bat_row['B_cost_e'],
            "Battery Cost P ($/kW)": bat_row['B_cost_p'],
            "Annualized PV Cost": ac_pv,
            "Annualized Battery Cost E": ac_e,
            "Annualized Battery Cost P": ac_p,

            # 5. Installed Component Costs
            "Installed PV Cost": installed_pv_cost,
            "Installed Battery Energy Cost": installed_battery_energy_cost,
            "Installed Battery Power Cost": installed_battery_power_cost,
            "Installed Battery Total Cost": installed_battery_cost,
    
            # 6. Generation
            "Annual PV Generation (kWh)": sum(ampl.getVariable("PV_gen").get((h, pv_row['PV_types'])).value() for h in data["hour_id"]),
            "Annual Grid Import (kWh)": sum(ampl.getVariable("Grid_buy").get(h).value() for h in data["hour_id"]),
            "Annual Grid Export (kWh)": sum(ampl.getVariable("Grid_sell").get(h).value() for h in data["hour_id"]),
    
            # 7. Generation Costs
            "Yearly Import Cost ($)": yearly_import_cost,
            "Off-Peak Import Cost ($)": monthly_df["OffPeak_Cost"].sum(),
            "On-Peak Import Cost ($)": monthly_df["OnPeak_Cost"].sum(),
            "Demand Charge ($)": monthly_df["Demand_Charge"].sum(),
            "Max Demand Charge ($)": monthly_df["Max_Demand_Charge"].sum(),
            "Total Sale Revenue": total_sale_revenue,
            #8. Emissions
            "Offset Price ($/tCO2)": offset_price,
            "Total Emissions Offset": Total_Emissions_Offset,
            "Gross Emissions": gross_emissions,
            "Net Emissions": gross_emissions - Total_Emissions_Offset,
            "Abatement Credit ($)": abatement_credit
        }


        for key, val in scalar_data.items():
            if val is None or (isinstance(val, float) and np.isnan(val)):
                print(f"⚠️ Warning: scalar '{key}' is undefined or NaN in run {run_id}")

       
        # Create DataFrames for scalar and hourly results
        scalar_df = pd.DataFrame([scalar_data])
        #small_scalar_df = pd.DataFrame(list( small_scalar_data()), columns=["Parameter", "Value"])    
        print(f"[{run_id}] Scalar DF preview:")
        print(scalar_df.head())
  
        write_results_to_excel(
            excel_path=excel_filename,
            pv_data=pv_data,
            bat_data=bat_data,
            override_params=override_params,
            run_id=run_id,
            sheet_name=hourly_sheet_name,
            scalar_df=scalar_df,
            hourly_df=hourly_df,
            monthly_df=monthly_df,
            sizing_limits=sizing_limits,
            monthly_sheet_name=monthly_sheet_name
        )
        log_simulation_complete(run_id, config_id_full)
        _insert_small_scalar_top(excel_filename, hourly_sheet_name, small_scalar_df)
        

            # ---- Write per-configuration hourly CSV with LAUNCH A_tot naming ----
     #   try:
    # Prefer the actual sweep file path if provided (so we can get the "first word"):
      #      sweep_source = override_params.get("sweep_file", excel_filename)

    # Force area to come from LAUNCH CSV, not optimized area
       #     overrides_for_name = {
        #        "A_tot": override_params.get("A_tot", None),   # <- this is the launch value you passed in
         #       "offset_price": offset_price,
          #  }

           # hourly_csv_name = make_hourly_csv_name(sweep_source, scalar_data, overrides_for_name)
           # hourly_csv_path = os.path.join(os.path.dirname(excel_filename), hourly_csv_name)
           # hourly_df.to_csv(hourly_csv_path, index=False)
           # print("💾 Hourly CSV:", hourly_csv_path)
        # except Exception as _e:
          #  print(f"⚠️ Hourly CSV write skipped: {str(_e)}")


        
# choose a source name for the first token
        sweep_source = override_params.get("sweep_file", excel_filename)  # use sweep file if provided, else excel name

# make sure we have the values the naming helper expects
        overrides_for_name = {
            "installed_area_m2": scalar_data["Installed Area (m^2)"],
            "itc_PV": itc_pv,
            "itc_Bat": itc_bat,
            "offset_price": offset_price,   # your scalar dict uses "Offset Price ($/tCO2)"; pass explicit here
        }

        hourly_csv_name = make_hourly_csv_name(sweep_source, scalar_data, overrides_for_name)
        hourly_csv_path = os.path.join(os.path.dirname(excel_filename), hourly_csv_name)



        #print(f"✅ [{run_id}] Completed PV: {pv_row['PV_types']} | Battery: {bat_row['BATTERY_TYPES']}")
        #print(f"    PV size: {P_pv_val:.2f} kW | Battery: {Total_Energy_val:.2f} kWh / {Total_Power_val:.2f} kW")
        #print(f"    Total Cost: ${ampl.getObjective('Total_Cost').value():,.2f}")
        #print(f"    Installed PV: ${installed_pv_cost:,.2f} | Battery: ${installed_battery_cost:,.2f}")
        return scalar_df, hourly_df


    except Exception as e:
        logging.error(f"[{run_id}] Error: {e}")
        return None, None
if __name__ == "__main__":
    import argparse
    import pandas as pd
    from microgrid.core.model_helpers import get_paths, create_ampl, reset_ampl_model
    # import your own run function from this module
    # from Completed_Loop_0_hour import run_single_pv

    parser = argparse.ArgumentParser(description="Single configuration microgrid run")
    parser.add_argument("--pv",  help="PV_types name (e.g., MonoSi_A)", required=True)
    parser.add_argument("--bat", help="BATTERY_TYPES name (e.g., LFP_2h)", required=True)
    parser.add_argument("--offset_price", type=float, default=0.0,
                        help="$/tCO2 for abatement monetization")
    parser.add_argument("--excel_out", default="single_run.xlsx",
                        help="Output Excel filename")
    parser.add_argument("--override", default="", help="CSV of overrides (optional)")
    args = parser.parse_args()

    paths = get_paths()
    ampl = create_ampl()
    reset_ampl_model(ampl, paths["model_path"])

    data    = pd.read_csv(paths["data_path"])
    pv_tab  = pd.read_csv(paths["pv_table_path"])
    bat_tab = pd.read_csv(paths["bat_table_path"])

    # choose rows by name
    pv_row  = pv_tab.loc[pv_tab["PV_types"] == args.pv].iloc[0]
    bat_row = bat_tab.loc[bat_tab["BATTERY_TYPES"] == args.bat].iloc[0]

    # build overrides (offset_price + anything else you want to test)
    override_params = {"offset_price": args.offset_price}
    if args.override:
        ov = pd.read_csv(args.override).to_dict(orient="records")[0]
        override_params.update(ov)

    # sized limits (optional precheck if your pipeline uses it)
    sizing_limits = {}

    # call the existing library function (no AMPL init inside it!)
    run_single_pv(
        run_id=f"{args.pv}_{args.bat}",
        ampl=ampl,
        data=data,
        pv_data=pv_tab,
        pv_row=pv_row,
        bat_data=bat_tab,
        bat_row=bat_row,
        excel_filename=args.excel_out,
        override_params=override_params,
        sizing_limits=sizing_limits,
    )
    print(f"Done. Wrote {args.excel_out}")
    