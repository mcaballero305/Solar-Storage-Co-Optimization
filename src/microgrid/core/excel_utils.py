
import pandas as pd
import numpy as np
import time
import os
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def _safe_load_workbook(path, tries=5, delay=0.4):
    last_err = None
    for _ in range(tries):
        try:
            return load_workbook(path)
        except Exception as e:
            last_err = e
            time.sleep(delay)
    raise last_err

def write_results_to_excel(
    excel_path: str,
    pv_data: pd.DataFrame,
    bat_data: pd.DataFrame,
    override_params: dict,
    run_id: str,
    sheet_name: str,
    scalar_df: pd.DataFrame,
    hourly_df: pd.DataFrame,
    monthly_df: pd.DataFrame = None,
    sizing_limits: dict = None,
    monthly_sheet_name: str = None
):
    """
    Writes scalar and hourly results, maintains an appended 'Summary' sheet,
    and optionally writes a monthly summary sheet.
    """
    enforce_limit_flag = bool(int(override_params.get("enforce_limit_flag", 0)))
    enable_sell = bool(int(override_params.get("enable_sell", 1)))

    initial_needed = not os.path.exists(excel_path)
    if initial_needed:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            pv_data.to_excel(writer, sheet_name='Initial', index=False)
            startrow = len(pv_data) + 2
            pd.DataFrame([['Battery Data']]).to_excel(writer, sheet_name='Initial', header=False, index=False, startrow=startrow)
            bat_data.to_excel(writer, sheet_name='Initial', index=False, startrow=startrow + 1)
            param_start = startrow + 1 + len(bat_data) + 2
            pd.DataFrame([['Sweep Parameters']]).to_excel(writer, sheet_name='Initial', header=False, index=False, startrow=param_start)
            pd.DataFrame([override_params]).to_excel(writer, sheet_name='Initial', index=False, startrow=param_start + 1)
            if sizing_limits:
                sizing_start = param_start + 1 + 1 + len(override_params)
                limits_df = pd.DataFrame(list(sizing_limits.items()), columns=['Metric', 'Value'])
                pd.DataFrame([['PV Sizing Limits']]).to_excel(writer, sheet_name='Initial', header=False, index=False, startrow=sizing_start)
                limits_df.to_excel(writer, sheet_name='Initial', index=False, startrow=sizing_start + 1)

    # Build final sheet name (Excel max 31 chars)
    run_sheet = (sheet_name or "Run").strip()[:31]

    # Write scalar + hourly on same sheet
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        scalar_df.to_excel(writer, sheet_name=run_sheet, index=False)
        hourly_df.to_excel(writer, sheet_name=run_sheet, startrow=len(scalar_df) + 3, index=False)

    print(f"✅ Wrote scalar + hourly combined for {run_id} to {excel_path} (Sheet: {run_sheet})")

    # Update Summary sheet (append-like behavior with column union)
    try:
        wb = _safe_load_workbook(excel_path)
        has_summary = 'Summary' in wb.sheetnames
    except Exception:
        has_summary = False

    if has_summary:
        existing = pd.read_excel(excel_path, sheet_name='Summary')
        # column union
        for col in scalar_df.columns:
            if col not in existing.columns:
                existing[col] = np.nan
        for col in existing.columns:
            if col not in scalar_df.columns:
                scalar_df[col] = np.nan
        summary_df = pd.concat([existing, scalar_df], ignore_index=True)
    else:
        summary_df = scalar_df.copy()

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    print(f"✅ Updated Summary sheet for {run_id}")

    # Monthly sheet as separate tab if provided
    if monthly_df is not None and monthly_sheet_name is not None:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            monthly_df.to_excel(writer, sheet_name=monthly_sheet_name[:31], index=False)
        print(f"✅ Wrote Monthly sheet for {run_id}")

    print(f"✅ Finished writing all results for {run_id}")

# === Suggested CSV naming for runs and summaries ===
# Use the following helpers in your sweep script after each solve:
#
#   from datetime import datetime
#
#   def make_run_slug(scalar_row: dict, fallback: str = "run") -> str:
#       parts = []
#       def add(name, key, fmt=lambda x: str(x)):
#           if key in scalar_row and pd.notna(scalar_row[key]):
#               parts.append(f"{name}{fmt(scalar_row[key])}".replace(" ", ""))
#       add("area", "Installed Area (m^2)", fmt=lambda x: f"{int(float(x))}m2")
#       add("pv", "PV Type")
#       add("bat", "Battery Type")
#       add("pvITC", "itc_PV", fmt=lambda x: f"{float(x):.2f}")
#       add("batITC", "itc_Bat", fmt=lambda x: f"{float(x):.2f}")
#       add("sell", "Enable Sell Flag", fmt=lambda x: "T" if bool(int(x)) else "F")
#       add("limit", "Utility Limit Applied?", fmt=lambda x: "T" if bool(int(x)) else "F")
#       if not parts:
#           return fallback
#       return "__".join(parts)
#
#   slug = make_run_slug(scalar_df.iloc[0].to_dict(), fallback=run_id)
#   per_run_csv = os.path.join(output_folder, f"{slug}__{datetime.now().strftime('%Y%m%d-%H%M%S')}.csv")
#   scalar_df.to_csv(per_run_csv, index=False)
